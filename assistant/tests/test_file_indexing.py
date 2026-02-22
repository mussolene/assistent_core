"""Tests for file indexing: chunking, extraction, file ref store, archives."""

import gzip
import tarfile
import zipfile
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import httpx
import pytest

from assistant.core import file_indexing as fi


def test_strip_html():
    assert fi._strip_html("<p>Hello</p>") == "Hello"
    assert fi._strip_html("<a href='x'>Link</a> text") == "Link text"
    assert "Hello" in fi._strip_html("<div>Hello <b>World</b></div>")


def test_is_archive():
    assert fi._is_archive(".zip", "") is True
    assert fi._is_archive("x.tar.gz", "") is True
    assert fi._is_archive("x.7z", "") is True
    assert fi._is_archive("x.rar", "") is True
    assert fi._is_archive("x.txt", "") is False
    assert fi._is_archive("x", "application/zip") is True


def test_extract_text_csv(tmp_path):
    f = tmp_path / "d.csv"
    f.write_text("a,b,c\n1,2,3", encoding="utf-8")
    out = fi._extract_text(f, "text/csv", "d.csv")
    assert "a" in out and "1" in out


def test_extract_text_html(tmp_path):
    f = tmp_path / "p.html"
    f.write_text("<html><body><p>Hello</p></body></html>", encoding="utf-8")
    out = fi._extract_text(f, "text/html", "p.html")
    assert "Hello" in out


def test_extract_text_md(tmp_path):
    f = tmp_path / "r.md"
    f.write_text("# Title\n\nBody text", encoding="utf-8")
    assert fi._extract_text(f, "", "r.md") == "# Title\n\nBody text"


def test_extract_text_xlsx(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Col1"
    ws["B1"] = "Col2"
    ws["A2"] = "a"
    ws["B2"] = "b"
    xlsx_path = tmp_path / "book.xlsx"
    wb.save(xlsx_path)
    out = fi._extract_text(
        xlsx_path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "book.xlsx"
    )
    assert "Col1" in out and "Col2" in out and "a" in out and "b" in out
    assert "Data" in out or "Лист" in out


def test_extract_content_from_file_txt(tmp_path):
    f = tmp_path / "t.txt"
    f.write_text("Plain text", encoding="utf-8")
    out = fi._extract_content_from_file(f, "text/plain", "t.txt")
    assert out == "Plain text"


def test_extract_content_from_file_zip_with_txt(tmp_path):
    zip_path = tmp_path / "a.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("inner.txt", "Content inside zip")
    out = fi._extract_content_from_file(zip_path, "application/zip", "a.zip")
    assert "Content inside zip" in out
    assert "inner.txt" in out


def test_extract_content_from_file_zip_path_traversal_ignored(tmp_path):
    zip_path = tmp_path / "bad.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("../../../etc/passwd", "skip")
    out = fi._extract_content_from_file(zip_path, "application/zip", "bad.zip")
    assert "skip" not in out


def test_extract_content_from_file_zip_skips_macosx_and_dirs(tmp_path):
    """Zip entries __MACOSX, .DS_Store and dirs (trailing /) are skipped."""
    zip_path = tmp_path / "mac.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("__MACOSX/._file", "ignore")
        zf.writestr("folder/.DS_Store", "ignore")
        zf.writestr("dir/", "")
        zf.writestr("ok.txt", "content")
    out = fi._extract_content_from_file(zip_path, "application/zip", "mac.zip")
    assert "content" in out
    assert "ignore" not in out


def test_extract_content_from_file_tar_with_txt(tmp_path):
    tar_path = tmp_path / "a.tar"
    with tarfile.open(tar_path, "w") as tf:
        inner = tmp_path / "inner.txt"
        inner.write_text("Content inside tar", encoding="utf-8")
        tf.add(inner, arcname="inner.txt")
    out = fi._extract_content_from_file(tar_path, "application/x-tar", "a.tar")
    assert "Content inside tar" in out
    assert "inner.txt" in out


def test_extract_content_from_file_tar_gz_with_txt(tmp_path):
    """Tar.gz is opened with tarfile and members extracted."""
    tar_gz = tmp_path / "a.tar.gz"
    inner = tmp_path / "inner.txt"
    inner.write_text("Inside tgz", encoding="utf-8")
    with tarfile.open(tar_gz, "w:gz") as tf:
        tf.add(inner, arcname="inner.txt")
    out = fi._extract_content_from_file(tar_gz, "application/gzip", "a.tar.gz")
    assert "Inside tgz" in out


def test_extract_content_from_file_single_gz(tmp_path):
    gz_path = tmp_path / "single.txt.gz"
    with gzip.open(gz_path, "wt", encoding="utf-8") as f:
        f.write("Gzipped text content")
    out = fi._extract_content_from_file(gz_path, "application/gzip", "single.txt.gz")
    assert "Gzipped text content" in out


def test_extract_content_from_file_max_files(tmp_path):
    """Zip with many files: extraction capped (zip iterates namelist()[:200])."""
    zip_path = tmp_path / "many.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        for i in range(300):
            zf.writestr(f"f{i}.txt", f"content {i}")
    out = fi._extract_content_from_file(zip_path, "application/zip", "many.zip")
    assert "content 0" in out
    assert "content 199" in out


def test_chunk_text_empty():
    assert fi._chunk_text("") == []
    assert fi._chunk_text("   ") == []


def test_chunk_text_short():
    text = "Hello world."
    assert len(fi._chunk_text(text)) == 1
    assert fi._chunk_text(text)[0] == text


def test_chunk_text_splits_with_overlap():
    text = "a" * 600
    chunks = fi._chunk_text(text, chunk_size=200, overlap=50)
    assert len(chunks) >= 2
    assert all(len(c) <= 200 for c in chunks)


def test_extract_text_txt(tmp_path):
    f = tmp_path / "f.txt"
    f.write_text("Hello\nWorld", encoding="utf-8")
    assert fi._extract_text(f, "text/plain", "f.txt") == "Hello\nWorld"


def test_extract_text_txt_read_error(tmp_path):
    """_extract_text for txt: read error -> returns ''."""
    f = tmp_path / "f.txt"
    f.write_text("x", encoding="utf-8")
    with patch.object(Path, "read_text", side_effect=OSError("Permission denied")):
        out = fi._extract_text(f, "text/plain", "f.txt")
    assert out == ""


def test_extract_text_image_returns_placeholder():
    # path can be any path; mime image -> placeholder
    p = Path("/nonexistent")
    assert "изображение" in fi._extract_text(p, "image/jpeg", "x.jpg")


def test_extract_text_pdf_with_pypdf(tmp_path):
    pytest.importorskip("pypdf")
    from pypdf import PdfWriter

    pdf_path = tmp_path / "t.pdf"
    w = PdfWriter()
    w.add_blank_page(72, 72)
    w.write(pdf_path)
    out = fi._extract_text(pdf_path, "application/pdf", "t.pdf")
    assert isinstance(out, str)


def test_extract_text_docx_with_docx(tmp_path):
    pytest.importorskip("docx")
    from docx import Document

    doc = Document()
    doc.add_paragraph("Hello from docx")
    doc_path = tmp_path / "t.docx"
    doc.save(doc_path)
    out = fi._extract_text(
        doc_path,
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "t.docx",
    )
    assert "Hello" in out


def test_extract_text_unknown_suffix_returns_empty(tmp_path):
    (tmp_path / "x.xyz").write_text("data")
    out = fi._extract_text(tmp_path / "x.xyz", "", "x.xyz")
    assert out == ""


def test_extract_text_pdf_import_error_returns_empty(tmp_path):
    """_extract_text for PDF when pypdf is not installed returns ''."""
    (tmp_path / "x.pdf").write_bytes(b"dummy")
    import sys

    class FakePypdf:
        def __getattr__(self, name):
            raise ImportError("No module named 'pypdf'")

    with patch.dict(sys.modules, {"pypdf": FakePypdf()}):
        out = fi._extract_text(tmp_path / "x.pdf", "application/pdf", "x.pdf")
    assert out == ""


def test_extract_text_pdf_exception_returns_empty(tmp_path):
    """_extract_text for PDF when extraction raises returns ''."""
    pytest.importorskip("pypdf")
    (tmp_path / "x.pdf").write_bytes(b"dummy")
    with patch("pypdf.PdfReader", side_effect=RuntimeError("corrupt pdf")):
        out = fi._extract_text(tmp_path / "x.pdf", "application/pdf", "x.pdf")
    assert out == ""


def test_extract_text_csv_exception_returns_empty(tmp_path):
    """_extract_text for CSV when read fails returns ''."""
    p = tmp_path / "x.csv"
    p.write_text("a,b", encoding="utf-8")
    with patch("builtins.open", side_effect=OSError("Permission denied")):
        out = fi._extract_text(p, "text/csv", "x.csv")
    assert out == ""


def test_extract_text_html_exception_returns_empty(tmp_path):
    """_extract_text for HTML when read fails returns ''."""
    p = tmp_path / "x.html"
    p.write_text("<p>Hi</p>", encoding="utf-8")
    with patch.object(Path, "read_text", side_effect=OSError("Permission denied")):
        out = fi._extract_text(p, "text/html", "x.html")
    assert out == ""


def test_extract_text_md_exception_returns_empty(tmp_path):
    """_extract_text for MD when read fails returns ''."""
    p = tmp_path / "x.md"
    p.write_text("# Hi", encoding="utf-8")
    with patch.object(Path, "read_text", side_effect=OSError("Permission denied")):
        out = fi._extract_text(p, "", "x.md")
    assert out == ""


def test_extract_text_xlsx_import_error_returns_empty(tmp_path):
    """_extract_text for XLSX when openpyxl is not installed returns ''."""
    (tmp_path / "x.xlsx").write_bytes(b"dummy")
    import sys

    class FakeOpenpyxl:
        def __getattr__(self, name):
            raise ImportError("No module named 'openpyxl'")

    with patch.dict(sys.modules, {"openpyxl": FakeOpenpyxl()}):
        out = fi._extract_text(
            tmp_path / "x.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
            "x.xlsx",
        )
    assert out == ""


@pytest.mark.asyncio
async def test_index_telegram_attachments_empty():
    ref_ids, text = await fi.index_telegram_attachments(
        "redis://localhost:6379/0",
        None,
        "u1",
        "c1",
        [],
        "",
    )
    assert ref_ids == []
    assert text == ""


@pytest.mark.asyncio
async def test_index_telegram_attachments_no_token():
    ref_ids, text = await fi.index_telegram_attachments(
        "redis://localhost:6379/0",
        None,
        "u1",
        "c1",
        [{"file_id": "x", "filename": "a.txt", "source": "telegram"}],
        "",
    )
    assert ref_ids == []
    assert text == ""


@pytest.mark.asyncio
async def test_index_telegram_attachments_skip_non_telegram():
    memory = MagicMock()
    memory.add_to_vector = AsyncMock()
    ref_ids, text = await fi.index_telegram_attachments(
        "redis://localhost:6379/0",
        memory,
        "u1",
        "c1",
        [{"file_id": "x", "filename": "a.txt", "source": "email"}],
        "token",
    )
    assert ref_ids == []
    assert text == ""
    memory.add_to_vector.assert_not_called()


@pytest.mark.asyncio
async def test_index_telegram_attachments_skip_no_file_id():
    memory = MagicMock()
    ref_ids, text = await fi.index_telegram_attachments(
        "redis://localhost:6379/0",
        memory,
        "u1",
        "c1",
        [{"filename": "a.txt", "source": "telegram"}],
        "token",
    )
    assert ref_ids == []
    assert text == ""


@pytest.mark.asyncio
async def test_index_telegram_attachments_getfile_fails():
    memory = MagicMock()
    mock_resp = MagicMock()
    mock_resp.json.return_value = {"ok": False, "description": "Bad Request"}
    with patch("assistant.core.file_indexing.httpx.AsyncClient") as ac:
        instance = MagicMock()
        instance.__aenter__ = AsyncMock(return_value=instance)
        instance.__aexit__ = AsyncMock(return_value=None)
        instance.get = AsyncMock(return_value=mock_resp)
        ac.return_value = instance
        ref_ids, text = await fi.index_telegram_attachments(
            "redis://localhost:6379/0",
            memory,
            "u1",
            "c1",
            [{"file_id": "f1", "filename": "a.txt", "source": "telegram"}],
            "token",
        )
    assert ref_ids == []
    assert text == ""


@pytest.mark.asyncio
async def test_index_telegram_attachments_getfile_ok_but_no_file_path_skips():
    """getFile returns ok True but result has no file_path -> attachment skipped, ref_ids empty."""
    memory = MagicMock()
    get_file_resp = MagicMock()
    get_file_resp.json.return_value = {"ok": True, "result": {}}
    with patch("assistant.core.file_indexing.httpx.AsyncClient") as ac:
        instance = MagicMock()
        instance.__aenter__ = AsyncMock(return_value=instance)
        instance.__aexit__ = AsyncMock(return_value=None)
        instance.get = AsyncMock(return_value=get_file_resp)
        ac.return_value = instance
        ref_ids, text = await fi.index_telegram_attachments(
            "redis://localhost:6379/0",
            memory,
            "u1",
            "c1",
            [{"file_id": "f1", "filename": "a.txt", "source": "telegram"}],
            "token",
        )
    assert ref_ids == []
    assert text == ""
    memory.add_to_vector.assert_not_called()


@pytest.mark.asyncio
async def test_index_telegram_attachments_download_raises_skips_attachment():
    """When download raises (e.g. HTTPStatusError), exception is caught, ref_ids stay empty for that attachment."""
    memory = MagicMock()
    memory.add_to_vector = AsyncMock()
    get_file_resp = MagicMock()
    get_file_resp.json.return_value = {"ok": True, "result": {"file_path": "documents/f1.txt"}}
    download_resp = MagicMock()
    download_resp.raise_for_status = MagicMock(
        side_effect=httpx.HTTPStatusError("500", request=MagicMock(), response=MagicMock())
    )
    get_calls = []

    async def fake_get(url, **kwargs):
        get_calls.append(url)
        if "getFile" in url:
            return get_file_resp
        return download_resp

    with patch("assistant.core.file_indexing.httpx.AsyncClient") as ac:
        instance = MagicMock()
        instance.__aenter__ = AsyncMock(return_value=instance)
        instance.__aexit__ = AsyncMock(return_value=None)
        instance.get = AsyncMock(side_effect=fake_get)
        ac.return_value = instance
        ref_ids, text = await fi.index_telegram_attachments(
            "redis://localhost:6379/0",
            memory,
            "u1",
            "c1",
            [{"file_id": "f1", "filename": "a.txt", "source": "telegram"}],
            "token",
        )
    assert ref_ids == []
    assert text == ""
    memory.add_to_vector.assert_not_called()


@pytest.mark.asyncio
async def test_index_telegram_attachments_success():
    memory = MagicMock()
    memory.add_to_vector = AsyncMock()
    get_file_resp = MagicMock()
    get_file_resp.json.return_value = {"ok": True, "result": {"file_path": "documents/file.txt"}}
    get_file_resp.raise_for_status = MagicMock()
    download_resp = MagicMock()
    download_resp.content = b"Hello from file"
    download_resp.raise_for_status = MagicMock()

    async def fake_get(url, **kwargs):
        if "getFile" in url:
            return get_file_resp
        return download_resp

    with patch("assistant.core.file_indexing.httpx.AsyncClient") as ac:
        instance = MagicMock()
        instance.__aenter__ = AsyncMock(return_value=instance)
        instance.__aexit__ = AsyncMock(return_value=None)
        instance.get = AsyncMock(side_effect=fake_get)
        ac.return_value = instance
        with patch("assistant.core.file_indexing._save_file_ref_sync"):
            ref_ids, text = await fi.index_telegram_attachments(
                "redis://localhost:6379/0",
                memory,
                "u1",
                "c1",
                [{"file_id": "f1", "filename": "file.txt", "source": "telegram"}],
                "bot_token",
            )
    assert len(ref_ids) == 1
    assert "Hello from file" in text
    assert memory.add_to_vector.call_count >= 1


def test_get_file_ref_missing():
    with patch("assistant.core.file_indexing._get_file_ref_sync", return_value=None):
        assert fi.get_file_ref("redis://localhost/0", "ref1") is None


def test_get_file_ref_found():
    with patch(
        "assistant.core.file_indexing._get_file_ref_sync",
        return_value={"file_id": "f1", "filename": "doc.pdf"},
    ):
        out = fi.get_file_ref("redis://localhost/0", "ref1")
        assert out == {"file_id": "f1", "filename": "doc.pdf"}


def test_get_file_ref_sync_invalid_json_returns_none():
    """_get_file_ref_sync returns None when stored value is not valid JSON."""
    mock_client = MagicMock()
    mock_client.get = MagicMock(return_value="{invalid json")
    with patch("redis.from_url", return_value=mock_client):
        out = fi._get_file_ref_sync("redis://localhost/0", "ref1")
    assert out is None


def test_list_file_refs_empty():
    with patch("assistant.core.file_indexing._list_file_refs_sync", return_value=[]):
        assert fi.list_file_refs("redis://localhost/0", "u1") == []


def test_list_file_refs_with_refs():
    with (
        patch(
            "assistant.core.file_indexing._list_file_refs_sync",
            return_value=["r1", "r2"],
        ),
        patch(
            "assistant.core.file_indexing._get_file_ref_sync",
            side_effect=lambda _u, rid: {"filename": f"f_{rid}.txt"},
        ),
    ):
        out = fi.list_file_refs("redis://localhost/0", "u1")
        assert len(out) == 2
        assert out[0]["file_ref_id"] == "r1" and out[0]["filename"] == "f_r1.txt"
        assert out[1]["file_ref_id"] == "r2" and out[1]["filename"] == "f_r2.txt"


def test_list_file_refs_skips_ref_when_get_returns_none():
    """list_file_refs skips refs for which _get_file_ref_sync returns None."""
    with (
        patch(
            "assistant.core.file_indexing._list_file_refs_sync",
            return_value=["r1", "r2"],
        ),
        patch(
            "assistant.core.file_indexing._get_file_ref_sync",
            side_effect=lambda _u, rid: {"filename": "a.txt"} if rid == "r1" else None,
        ),
    ):
        out = fi.list_file_refs("redis://localhost/0", "u1")
        assert len(out) == 1
        assert out[0]["file_ref_id"] == "r1" and out[0]["filename"] == "a.txt"


def test_extract_content_from_file_respects_file_count_limit(tmp_path):
    """When file_count['n'] already >= MAX_ARCHIVE_FILES, returns '' without reading."""
    f = tmp_path / "t.txt"
    f.write_text("hello", encoding="utf-8")
    file_count = {"n": fi.MAX_ARCHIVE_FILES}
    out = fi._extract_content_from_file(f, "text/plain", "t.txt", 0, file_count)
    assert out == ""


def test_extract_content_from_file_at_max_depth_treats_archive_as_file(tmp_path):
    """When depth >= MAX_ARCHIVE_DEPTH, archive is not unpacked, _extract_text is used (zip -> '')."""
    zip_path = tmp_path / "a.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("inner.txt", "content")
    file_count = {"n": 0}
    out = fi._extract_content_from_file(
        zip_path, "application/zip", "a.zip", depth=fi.MAX_ARCHIVE_DEPTH, file_count=file_count
    )
    assert out == ""  # .zip with unknown content type in _extract_text returns ""


def test_strip_html_fallback_on_parser_error():
    """_strip_html falls back to regex when parser.feed raises."""
    with patch("html.parser.HTMLParser.feed", side_effect=ValueError("bad")):
        out = fi._strip_html("<p>Hi</p>")
    assert "Hi" in out
    assert "<" not in out or "p" not in out


def test_extract_from_archive_outer_exception_returns_joined_parts(tmp_path):
    """When archive extraction raises, _extract_from_archive returns joined parts (or empty)."""
    bad_zip = tmp_path / "x.zip"
    bad_zip.write_bytes(b"not a zip")
    file_count = {"n": 0}
    out = fi._extract_from_archive(bad_zip, "x.zip", 0, file_count)
    assert out == ""


def test_extract_from_archive_7z_import_error(tmp_path):
    """When py7zr is not installed, 7z extraction is skipped (ImportError)."""
    try:
        import py7zr  # noqa: F401

        pytest.skip("py7zr installed, cannot test ImportError path")
    except ImportError:
        pass
    seven_z = tmp_path / "x.7z"
    seven_z.write_bytes(b"dummy")
    file_count = {"n": 0}
    out = fi._extract_from_archive(seven_z, "x.7z", 0, file_count)
    assert out == ""
