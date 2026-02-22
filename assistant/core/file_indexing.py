"""Индексация вложений: извлечение текста из файлов и архивов → чанки → векторная память; ссылки в Redis."""

from __future__ import annotations

import csv
import gzip
import html.parser
import json
import logging
import os
import re
import shutil
import tarfile
import tempfile
import uuid
import zipfile
from pathlib import Path
from typing import Any, MutableMapping

import httpx

logger = logging.getLogger(__name__)

FILE_REF_PREFIX = "file_ref:"
FILE_REF_USER_PREFIX = "file_ref_user:"
CHUNK_SIZE = 500
CHUNK_OVERLAP = 50
MAX_ARCHIVE_DEPTH = 3
MAX_ARCHIVE_FILES = 500


def _extract_text(path: Path, mime_type: str, filename: str) -> str:
    """Извлечь текст из файла. Поддержка: txt, pdf, docx, csv, xlsx, html, md, изображения (OCR)."""
    suffix = (filename or path.name).lower()
    if suffix.endswith(".txt") or "text/plain" in (mime_type or ""):
        try:
            return path.read_text(encoding="utf-8", errors="replace")
        except Exception as e:
            logger.warning("Read text file %s: %s", path, e)
            return ""
    if suffix.endswith(".pdf") or "pdf" in (mime_type or ""):
        try:
            from pypdf import PdfReader

            reader = PdfReader(str(path))
            return "\n".join(p.extract_text() or "" for p in reader.pages)
        except ImportError:
            logger.debug("pypdf not installed, skip PDF extraction")
            return ""
        except Exception as e:
            logger.warning("PDF extraction %s: %s", path, e)
            return ""
    if suffix.endswith(".docx") or "wordprocessingml" in (mime_type or ""):
        try:
            from docx import Document

            doc = Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs)
        except ImportError:
            logger.debug("python-docx not installed, skip DOCX extraction")
            return ""
        except Exception as e:
            logger.warning("DOCX extraction %s: %s", path, e)
            return ""
    if suffix.endswith(".csv") or "csv" in (mime_type or ""):
        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                rows = list(reader)
            return "\n".join("\t".join(cell for cell in row) for row in rows[:2000])
        except Exception as e:
            logger.warning("CSV extraction %s: %s", path, e)
            return ""
    if suffix.endswith(".xlsx") or "spreadsheet" in (mime_type or ""):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            parts = []
            for sheet in wb.worksheets[:20]:
                parts.append(f"[Лист: {sheet.title}]")
                for row in sheet.iter_rows(max_row=2000, values_only=True):
                    parts.append("\t".join(str(c) if c is not None else "" for c in row))
            return "\n".join(parts)
        except ImportError:
            logger.debug("openpyxl not installed, skip XLSX")
            return ""
        except Exception as e:
            logger.warning("XLSX extraction %s: %s", path, e)
            return ""
    if suffix.endswith(".html") or suffix.endswith(".htm") or "html" in (mime_type or ""):
        try:
            raw = path.read_text(encoding="utf-8", errors="replace")
            return _strip_html(raw)
        except Exception as e:
            logger.warning("HTML extraction %s: %s", path, e)
            return ""
    if suffix.endswith(".md") or suffix.endswith(".markdown"):
        try:
            return path.read_text(encoding="utf-8", errors="replace")
        except Exception as e:
            logger.warning("Markdown read %s: %s", path, e)
            return ""
    if "image/" in (mime_type or ""):
        # OCR: извлечь текст из изображения (скриншоты, фото текста)
        try:
            import pytesseract
            from PIL import Image

            img = Image.open(path)
            # Поддержка русского и английского; при отсутствии rus данные в eng
            text = pytesseract.image_to_string(img, lang="rus+eng")
            if text and text.strip():
                return text.strip()
        except ImportError:
            logger.debug("pytesseract/pillow not installed, skip image OCR")
        except Exception as e:
            # TesseractNotFoundError или ошибка распознавания
            logger.debug("Image OCR %s: %s", path, e)
        return " [изображение] "
    return ""


def _strip_html(html_str: str) -> str:
    """Удалить теги HTML, оставить текст."""
    class _TextExtractor(html.parser.HTMLParser):
        def __init__(self) -> None:
            super().__init__()
            self.text: list[str] = []

        def handle_data(self, data: str) -> None:
            self.text.append(data)

    try:
        parser = _TextExtractor()
        parser.feed(html_str)
        return re.sub(r"\s+", " ", " ".join(parser.text)).strip()
    except Exception:
        return re.sub(r"<[^>]+>", " ", html_str)


def _is_archive(suffix: str, mime_type: str) -> bool:
    """Является ли файл архивом (zip, tar, gz, 7z, rar)."""
    archive_suffixes = (".zip", ".tar", ".tgz", ".tar.gz", ".tar.bz2", ".tbz2", ".gz", ".7z", ".rar")
    return any(suffix.endswith(s) for s in archive_suffixes) or "zip" in (mime_type or "")


def _extract_content_from_file(
    path: Path,
    mime_type: str,
    filename: str,
    depth: int = 0,
    file_count: MutableMapping[str, int] | None = None,
) -> str:
    """
    Извлечь весь текст из файла или архива (рекурсивно).
    Ограничения: глубина вложенности архивов MAX_ARCHIVE_DEPTH, всего файлов MAX_ARCHIVE_FILES.
    """
    if file_count is None:
        file_count = {"n": 0}
    if file_count["n"] >= MAX_ARCHIVE_FILES:
        return ""
    suffix = (filename or path.name).lower()
    if depth < MAX_ARCHIVE_DEPTH and _is_archive(suffix, mime_type):
        return _extract_from_archive(path, filename, depth, file_count)
    file_count["n"] += 1
    return _extract_text(path, mime_type, filename)


def _extract_from_archive(
    path: Path,
    archive_name: str,
    depth: int,
    file_count: MutableMapping[str, int],
) -> str:
    """Распаковать архив и извлечь текст из всех вложенных файлов."""
    parts: list[str] = []
    suffix = (archive_name or path.name).lower()
    try:
        if suffix.endswith(".zip"):
            with zipfile.ZipFile(path, "r") as zf:
                for name in zf.namelist()[:200]:
                    if file_count["n"] >= MAX_ARCHIVE_FILES:
                        break
                    if name.endswith("/") or "__MACOSX" in name or ".DS_Store" in name:
                        continue
                    safe_name = Path(name).name
                    if not safe_name or ".." in name:
                        continue
                    try:
                        data = zf.read(name)
                        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(safe_name).suffix) as tmp:
                            tmp.write(data)
                            tmp_path = Path(tmp.name)
                        try:
                            text = _extract_content_from_file(
                                tmp_path, "", safe_name, depth + 1, file_count
                            )
                            if text.strip():
                                parts.append(f"[{name}]\n{text}")
                        finally:
                            if tmp_path.exists():
                                tmp_path.unlink(missing_ok=True)
                    except Exception as e:
                        logger.debug("Zip member %s: %s", name, e)
        elif suffix.endswith(".tar") or suffix.endswith(".tar.gz") or suffix.endswith(".tgz") or suffix.endswith(".tar.bz2") or suffix.endswith(".tbz2"):
            with tarfile.open(path, "r:*") as tf:
                for member in tf.getmembers()[:200]:
                    if file_count["n"] >= MAX_ARCHIVE_FILES:
                        break
                    if not member.isfile() or ".." in member.name or "__MACOSX" in member.name:
                        continue
                    safe_name = Path(member.name).name
                    if not safe_name:
                        continue
                    try:
                        f = tf.extractfile(member)
                        if f is None:
                            continue
                        data = f.read()
                        f.close()
                        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(safe_name).suffix) as tmp:
                            tmp.write(data)
                            tmp_path = Path(tmp.name)
                        try:
                            text = _extract_content_from_file(
                                tmp_path, "", safe_name, depth + 1, file_count
                            )
                            if text.strip():
                                parts.append(f"[{member.name}]\n{text}")
                        finally:
                            if tmp_path.exists():
                                tmp_path.unlink(missing_ok=True)
                    except Exception as e:
                        logger.debug("Tar member %s: %s", member.name, e)
        elif suffix.endswith(".gz") and not suffix.endswith(".tar.gz"):
            try:
                with gzip.open(path, "rt", encoding="utf-8", errors="replace") as f:
                    return f.read()[:500_000]
            except Exception as e:
                logger.debug("Gzip read %s: %s", path, e)
                return ""
        elif suffix.endswith(".7z"):
            try:
                import py7zr
                with py7zr.SevenZipFile(path, "r") as zf:
                    tmpdir = Path(tempfile.mkdtemp())
                    try:
                        zf.extractall(tmpdir)
                        for fpath in tmpdir.rglob("*")[:200]:
                            if file_count["n"] >= MAX_ARCHIVE_FILES:
                                break
                            if fpath.is_file():
                                rel = fpath.relative_to(tmpdir)
                                text = _extract_content_from_file(
                                    fpath, "", str(rel), depth + 1, file_count
                                )
                                if text.strip():
                                    parts.append(f"[{rel}]\n{text}")
                    finally:
                        shutil.rmtree(tmpdir, ignore_errors=True)
            except ImportError:
                logger.debug("py7zr not installed, skip 7z")
            except Exception as e:
                logger.warning("7z extraction %s: %s", path, e)
        elif suffix.endswith(".rar"):
            try:
                from rarfile import RarFile
                with RarFile(path, "r") as rf:
                    tmpdir = Path(tempfile.mkdtemp())
                    try:
                        rf.extractall(tmpdir)
                        for fpath in tmpdir.rglob("*")[:200]:
                            if file_count["n"] >= MAX_ARCHIVE_FILES:
                                break
                            if fpath.is_file():
                                rel = fpath.relative_to(tmpdir)
                                text = _extract_content_from_file(
                                    fpath, "", str(rel), depth + 1, file_count
                                )
                                if text.strip():
                                    parts.append(f"[{rel}]\n{text}")
                    finally:
                        shutil.rmtree(tmpdir, ignore_errors=True)
            except ImportError:
                logger.debug("rarfile not installed, skip rar")
            except Exception as e:
                logger.warning("RAR extraction %s: %s", path, e)
    except Exception as e:
        logger.warning("Archive extraction %s: %s", path, e)
    return "\n\n".join(parts)


def _chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    """Разбить текст на чанки с перекрытием."""
    if not text or not text.strip():
        return []
    chunks = []
    start = 0
    text = text.strip()
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        if chunk.strip():
            chunks.append(chunk.strip())
        start = end - overlap if overlap < chunk_size else end
    return chunks


def _save_file_ref_sync(redis_url: str, ref_id: str, user_id: str, data: dict[str, Any]) -> None:
    import redis

    client = redis.from_url(redis_url, decode_responses=True)
    client.set(FILE_REF_PREFIX + ref_id, json.dumps(data, ensure_ascii=False))
    client.sadd(FILE_REF_USER_PREFIX + user_id, ref_id)


def _get_file_ref_sync(redis_url: str, ref_id: str) -> dict[str, Any] | None:
    import redis

    client = redis.from_url(redis_url, decode_responses=True)
    raw = client.get(FILE_REF_PREFIX + ref_id)
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        return None


def _list_file_refs_sync(redis_url: str, user_id: str) -> list[str]:
    import redis

    client = redis.from_url(redis_url, decode_responses=True)
    refs = client.smembers(FILE_REF_USER_PREFIX + user_id)
    return list(refs) if refs else []


# Максимум символов извлечённого текста для передачи в оркестратор (summary)
EXTRACTED_TEXT_CAP = 6000


async def index_telegram_attachments(
    redis_url: str,
    memory: Any,
    user_id: str,
    chat_id: str,
    attachments: list[dict[str, Any]],
    bot_token: str,
) -> tuple[list[str], str]:
    """
    Скачать вложения из Telegram, извлечь текст, положить чанки в векторную память,
    сохранить ссылки на файлы в Redis (file_id для последующей отправки по запросу).
    Возвращает (список file_ref_id, извлечённый текст до EXTRACTED_TEXT_CAP символов для summary).
    """
    if not bot_token or not attachments:
        return [], ""
    base_url = f"https://api.telegram.org/bot{bot_token}"
    ref_ids: list[str] = []
    extracted_parts: list[str] = []
    for att in attachments:
        if att.get("source") != "telegram":
            continue
        file_id = att.get("file_id")
        filename = att.get("filename") or "file"
        mime_type = att.get("mime_type") or ""
        if not file_id:
            continue
        ref_id = str(uuid.uuid4())[:12]
        try:
            async with httpx.AsyncClient() as client:
                r = await client.get(f"{base_url}/getFile", params={"file_id": file_id}, timeout=10.0)
            data = r.json()
            if not data.get("ok"):
                logger.warning("Telegram getFile failed: %s", data)
                continue
            file_path = data.get("result", {}).get("file_path")
            if not file_path:
                continue
            download_url = f"https://api.telegram.org/file/bot{bot_token}/{file_path}"
            with tempfile.NamedTemporaryFile(delete=False, suffix=Path(filename).suffix) as tmp:
                tmp_path = Path(tmp.name)
            try:
                async with httpx.AsyncClient() as client:
                    resp = await client.get(download_url, timeout=30.0)
                resp.raise_for_status()
                tmp_path.write_bytes(resp.content)
                text = _extract_content_from_file(tmp_path, mime_type, filename)
                if text.strip():
                    extracted_parts.append(f"[{filename}]\n{text}")
                chunks = _chunk_text(text)
                for i, chunk in enumerate(chunks):
                    await memory.add_to_vector(
                        user_id,
                        chunk,
                        metadata={
                            "source": "file",
                            "file_ref_id": ref_id,
                            "filename": filename,
                            "chunk_index": i,
                        },
                    )
                _save_file_ref_sync(
                    redis_url,
                    ref_id,
                    user_id,
                    {
                        "file_id": file_id,
                        "chat_id": chat_id,
                        "user_id": user_id,
                        "filename": filename,
                        "source": "telegram",
                    },
                )
                ref_ids.append(ref_id)
            finally:
                if tmp_path.exists():
                    try:
                        tmp_path.unlink()
                    except OSError:
                        pass
        except Exception as e:
            logger.exception("Index attachment %s: %s", filename, e)
    combined = "\n\n".join(extracted_parts)
    if len(combined) > EXTRACTED_TEXT_CAP:
        combined = combined[: EXTRACTED_TEXT_CAP] + "\n\n[...]"
    return ref_ids, combined


def get_file_ref(redis_url: str, ref_id: str) -> dict[str, Any] | None:
    """Получить ссылку на файл по ref_id (для отправки в чат по file_id)."""
    return _get_file_ref_sync(redis_url, ref_id)


def list_file_refs(redis_url: str, user_id: str) -> list[dict[str, Any]]:
    """Список сохранённых ссылок на файлы пользователя (filename, ref_id)."""
    ref_ids = _list_file_refs_sync(redis_url, user_id)
    result = []
    for rid in ref_ids:
        ref = _get_file_ref_sync(redis_url, rid)
        if ref:
            result.append({"file_ref_id": rid, "filename": ref.get("filename") or rid})
    return result
